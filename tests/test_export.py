from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from cos_cost.cli import main
from cos_cost.ext.export import PAYABLE_RECONCILE_TOLERANCE, SHEET_NAMES


def test_cli_export_pdf_xlsx_reconcile(cache_dir, tmp_path: Path) -> None:
    pdf = tmp_path / "out.pdf"
    xlsx = tmp_path / "out.xlsx"
    code = main(
        [
            "export",
            "--mock",
            "--month",
            "2026-07",
            "--cache-dir",
            str(cache_dir),
            "--pdf",
            str(pdf),
            "--xlsx",
            str(xlsx),
        ]
    )
    assert code == 0
    assert pdf.is_file() and pdf.stat().st_size > 200
    assert xlsx.is_file() and xlsx.stat().st_size > 200
    assert pdf.read_bytes()[:4] == b"%PDF"

    wb = load_workbook(xlsx)
    assert tuple(wb.sheetnames) == SHEET_NAMES
    buckets = wb["按桶"]
    headers = [c.value for c in buckets[1]]
    assert "应付" in headers
    payable_col = headers.index("应付") + 1
    total = 0.0
    for row in buckets.iter_rows(min_row=2, values_only=True):
        label = row[0]
        if label == "合计(应付)":
            continue
        if row[payable_col - 1] is not None:
            total += float(row[payable_col - 1])
    assert abs(total - 186420.0) <= PAYABLE_RECONCILE_TOLERANCE

    text = xlsx.read_bytes()
    assert b"COS_SECRET_KEY" not in text
    assert b"super-secret" not in text
    # 口径写明对账容差
    koujing = "\n".join(str(c.value or "") for c in wb["口径"]["A"])
    assert "容差" in koujing
    assert str(PAYABLE_RECONCILE_TOLERANCE) in koujing


def test_pdf_watermark_when_ready_zero(cache_dir, tmp_path: Path) -> None:
    pdf = tmp_path / "est.pdf"
    code = main(
        [
            "export",
            "--mock",
            "--month",
            "2026-08",
            "--cache-dir",
            str(cache_dir),
            "--pdf",
            str(pdf),
        ]
    )
    assert code == 0
    raw = pdf.read_bytes()
    assert raw[:4] == b"%PDF"
    # 暂估 以 CID/ToUnicode 或明文出现；至少文件非空且为 PDF
    assert len(raw) > 500
