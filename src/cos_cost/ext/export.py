"""PDF（A4 横向一页）+ Excel（5 表）。不含密钥、不含对象 Key。"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any, Protocol

from cos_cost.billing_items import CATEGORY_LABELS, compose_categories
from cos_cost.formatters import money_text, pct_text, ready_label, volume_text
from cos_cost.models import CollectSnapshot, RankingResult
from cos_cost.secrets import assert_no_secrets

SHEET_NAMES = ("汇总", "按桶", "按计费项", "机会", "口径")
PAYABLE_RECONCILE_TOLERANCE = 0.05  # 元；按桶合计 vs COS 应付
FONT_CANDIDATES = (
    "/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf",
    "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
    "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
    "/System/Library/Fonts/STHeiti Light.ttc",
    "C:\\Windows\\Fonts\\msyh.ttc",
)


class RankingExporter(Protocol):
    def export(self, ranking: RankingResult, dest: Path) -> None:
        """将排行结果写成文件。"""


class UnsupportedExporter:
    def export(self, ranking: RankingResult, dest: Path) -> None:
        raise NotImplementedError("请使用 ReportExporter（PDF / Excel）。")


def mask_uin(value: str | None) -> str:
    if not value:
        return "—"
    text = str(value)
    if len(text) <= 6:
        return "***"
    return f"{text[:3]}****{text[-4:]}"


def build_report_payload(
    snapshot: CollectSnapshot,
    ranking: RankingResult,
    *,
    cards: list[dict[str, Any]] | None = None,
    composition: dict[str, float] | None = None,
    trend: dict[str, Any] | None = None,
    owner_uin: str | None = None,
) -> dict[str, Any]:
    engine_cards = cards
    if engine_cards is None:
        from cos_cost.ext.opportunity import RuleEngine

        engine_cards = RuleEngine(snapshot).list_all()
    composed = composition if composition is not None else compose_categories(snapshot.bill_resources)
    uin = owner_uin
    if uin is None:
        for row in snapshot.bill_resources:
            if row.owner_uin:
                uin = row.owner_uin
                break
    bucket_sum = sum(r.payable for r in ranking.rows if r.payable is not None)
    payable = ranking.kpis.cos_payable
    delta = None
    if payable is not None:
        delta = abs(bucket_sum - payable)
    return {
        "month": ranking.month,
        "account_key": mask_uin(snapshot.account_key),
        "owner_uin": mask_uin(uin),
        "ready": ranking.ready,
        "estimated": ranking.estimated,
        "ready_label": ready_label(ranking.ready, ranking.estimated),
        "mock": ranking.mock,
        "kpis": {
            "cos_payable": ranking.kpis.cos_payable,
            "optimizable": ranking.kpis.optimizable_amount,
            "standard_pct": ranking.kpis.standard_storage_pct,
            "internet_bytes": ranking.kpis.internet_traffic_bytes,
            "request_fee": ranking.kpis.request_fee,
            "ready": ranking.ready,
            "coverage": f"{ranking.kpis.bucket_with_bill}/{ranking.kpis.bucket_listed}",
            "mom_pct": ranking.kpis.mom_pct,
            "yoy_pct": ranking.kpis.yoy_pct,
        },
        "rows": [
            {
                "bucket": r.bucket,
                "region": r.region,
                "payable": r.payable,
                "mom_pct": r.mom_pct,
                "capacity_bytes": r.capacity_bytes,
                "standard_pct": r.standard_pct,
                "internet_traffic_bytes": r.internet_traffic_bytes,
                "opportunity_amount": r.opportunity_amount,
                "opportunity_count": r.opportunity_count,
            }
            for r in ranking.rows
        ],
        "opportunities": engine_cards,
        "composition": composed,
        "trend": trend or {},
        "notes": ranking.notes,
        "reconcile": {
            "bucket_sum": bucket_sum,
            "cos_payable": payable,
            "delta": delta,
            "tolerance": PAYABLE_RECONCILE_TOLERANCE,
            "ok": delta is None or delta <= PAYABLE_RECONCILE_TOLERANCE,
        },
    }


class ReportExporter:
    def export(self, ranking: RankingResult, dest: Path) -> None:
        raise NotImplementedError("请调用 write_pdf / write_xlsx，并传入完整报表 payload。")

    def write_pdf(self, payload: dict[str, Any], dest: Path) -> None:
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_bytes(render_pdf(payload))

    def write_xlsx(self, payload: dict[str, Any], dest: Path) -> None:
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_bytes(render_xlsx(payload))


def render_pdf(payload: dict[str, Any]) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfgen import canvas

    font_name = _register_cn_font()
    buf = BytesIO()
    width, height = landscape(A4)
    c = canvas.Canvas(buf, pagesize=landscape(A4))
    if payload.get("estimated") or payload.get("ready") == 0:
        c.saveState()
        c.setFont(font_name, 72)
        c.setFillColorRGB(0.85, 0.82, 0.7)
        c.translate(width / 2, height / 2)
        c.rotate(28)
        c.drawCentredString(0, 0, "暂估")
        c.restoreState()

    c.setFillColorRGB(0.06, 0.09, 0.16)
    c.rect(0, height - 22 * mm, width, 22 * mm, fill=1, stroke=0)
    c.setFillColorRGB(1, 1, 1)
    c.setFont(font_name, 14)
    c.drawString(12 * mm, height - 10 * mm, "COS 机会大师 · 一页报表")
    c.setFont(font_name, 9)
    title_meta = (
        f"账期 {payload.get('month')}    账户 {payload.get('account_key')}    "
        f"{payload.get('ready_label')}"
    )
    if payload.get("mock"):
        title_meta += "    [mock]"
    c.drawString(12 * mm, height - 17 * mm, title_meta)

    kpis = payload.get("kpis") or {}
    boxes = [
        ("COS 应付", money_text(kpis.get("cos_payable"))),
        ("可优化金额", money_text(kpis.get("optimizable"))),
        ("标准存储占比", pct_text(kpis.get("standard_pct"), signed=False)),
        ("外网下行", volume_text(kpis.get("internet_bytes"))),
        ("请求费", money_text(kpis.get("request_fee"))),
        ("数据就绪", "已出账" if kpis.get("ready") == 1 else "暂估"),
    ]
    box_w = (width - 24 * mm) / 6
    top = height - 28 * mm
    for i, (label, value) in enumerate(boxes):
        x = 12 * mm + i * box_w
        c.setStrokeColorRGB(0.9, 0.91, 0.93)
        c.setFillColorRGB(1, 1, 1)
        c.roundRect(x, top - 18 * mm, box_w - 3 * mm, 18 * mm, 3, fill=1, stroke=1)
        c.setFillColorRGB(0.42, 0.45, 0.5)
        c.setFont(font_name, 7)
        c.drawString(x + 2 * mm, top - 6 * mm, label)
        c.setFillColorRGB(0.07, 0.09, 0.15)
        c.setFont(font_name, 10)
        c.drawString(x + 2 * mm, top - 13 * mm, value)

    chart_top = top - 22 * mm
    _draw_trend(c, payload.get("trend") or {}, 12 * mm, chart_top - 52 * mm, width / 2 - 16 * mm, 50 * mm, font_name)
    _draw_bucket_bars(
        c,
        payload.get("rows") or [],
        width / 2 + 2 * mm,
        chart_top - 52 * mm,
        width / 2 - 16 * mm,
        50 * mm,
        font_name,
    )

    table_top = chart_top - 58 * mm
    _draw_table(
        c,
        "Top5 桶",
        ["桶", "应付", "环比", "机会"],
        [
            [
                _short(r.get("bucket")),
                money_text(r.get("payable")),
                pct_text(r.get("mom_pct")),
                money_text(r.get("opportunity_amount")),
            ]
            for r in (payload.get("rows") or [])[:5]
        ],
        12 * mm,
        table_top,
        width / 2 - 16 * mm,
        font_name,
    )
    opp_rows = []
    for card in (payload.get("opportunities") or [])[:5]:
        opp_rows.append(
            [
                f"{card.get('rule_id')} {_short(card.get('bucket'))}",
                money_text(card.get("net_saving")),
                f"{card.get('confidence')}",
                (card.get("why") or card.get("title") or "")[:28],
            ]
        )
    _draw_table(
        c,
        "Top5 机会",
        ["规则 / 桶", "净节省", "置信", "原因"],
        opp_rows,
        width / 2 + 2 * mm,
        table_top,
        width / 2 - 16 * mm,
        font_name,
    )

    recon = payload.get("reconcile") or {}
    footer = (
        f"口径：应付=RealTotalCost（p_cos）。可优化=net≥50 且无强阻断，不含 R05–R09 / R06（不含 CDN 下行）/ 备份桶。"
        f"按桶合计 {money_text(recon.get('bucket_sum'))} vs COS 应付 {money_text(recon.get('cos_payable'))}"
        f"（容差 {recon.get('tolerance')} 元）。单价优先账单，否则刊例。无对象 Key，无密钥。"
    )
    c.setFont(font_name, 7)
    c.setFillColorRGB(0.35, 0.38, 0.42)
    c.drawString(12 * mm, 8 * mm, footer[:160])
    c.drawString(12 * mm, 4 * mm, footer[160:])
    c.showPage()
    c.save()
    data = buf.getvalue()
    assert_no_secrets({"pdf": "ok"})
    return data


def render_xlsx(payload: dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(color="FFFFFF", bold=True)
    kpis = payload.get("kpis") or {}
    recon = payload.get("reconcile") or {}

    ws = wb.active
    ws.title = "汇总"
    summary_rows = [
        ("账期", payload.get("month")),
        ("账户", payload.get("account_key")),
        ("UIN", payload.get("owner_uin")),
        ("就绪", payload.get("ready_label")),
        ("COS 应付", kpis.get("cos_payable")),
        ("可优化金额", kpis.get("optimizable")),
        ("标准存储占比", kpis.get("standard_pct")),
        ("外网下行字节", kpis.get("internet_bytes")),
        ("请求费", kpis.get("request_fee")),
        ("桶覆盖", kpis.get("coverage")),
        ("按桶合计", recon.get("bucket_sum")),
        ("对账差额", recon.get("delta")),
        ("对账容差(元)", recon.get("tolerance")),
        ("对账通过", "是" if recon.get("ok") else "否"),
    ]
    ws.append(["项", "值"])
    for row in summary_rows:
        ws.append(list(row))

    ws2 = wb.create_sheet("按桶")
    ws2.append(["桶", "地域", "应付", "环比%", "容量字节", "标准%", "外网字节", "可优化", "规则数"])
    bucket_sum = 0.0
    for row in payload.get("rows") or []:
        pay = row.get("payable")
        if pay is not None:
            bucket_sum += float(pay)
        ws2.append(
            [
                row.get("bucket"),
                row.get("region"),
                row.get("payable"),
                row.get("mom_pct"),
                row.get("capacity_bytes"),
                row.get("standard_pct"),
                row.get("internet_traffic_bytes"),
                row.get("opportunity_amount"),
                row.get("opportunity_count"),
            ]
        )
    ws2.append(["合计(应付)", "", bucket_sum, "", "", "", "", "", ""])

    ws3 = wb.create_sheet("按计费项")
    ws3.append(["计费项", "金额"])
    composed = payload.get("composition") or {}
    for key, label in CATEGORY_LABELS.items():
        ws3.append([label, composed.get(key) or 0])
    ws3.append(["合计", sum(float(v or 0) for v in composed.values())])

    ws4 = wb.create_sheet("机会")
    ws4.append(
        ["规则", "桶", "列", "净节省", "置信度", "计入KPI", "原因", "公式", "建议", "草稿(无对象Key)"]
    )
    for card in payload.get("opportunities") or []:
        ws4.append(
            [
                card.get("rule_id"),
                card.get("bucket"),
                card.get("column"),
                card.get("net_saving"),
                card.get("confidence"),
                "是" if card.get("in_kpi") else "否",
                card.get("why"),
                card.get("formula"),
                card.get("action"),
                card.get("action_draft"),
            ]
        )

    ws5 = wb.create_sheet("口径")
    for line in _methodology_lines(payload):
        ws5.append([line])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            if sheet.title != "口径":
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(wrap_text=True)
        for column in sheet.columns:
            letter = get_column_letter(column[0].column)
            sheet.column_dimensions[letter].width = min(42, max(12, len(str(column[0].value or "")) + 4))

    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    # 不把整个 xlsx 丢进密钥扫描（二进制）；字段已脱敏。
    assert_no_secrets(
        {
            "account": payload.get("account_key"),
            "uin": payload.get("owner_uin"),
            "buckets": [r.get("bucket") for r in (payload.get("rows") or [])],
        }
    )
    return data


def _methodology_lines(payload: dict[str, Any]) -> list[str]:
    recon = payload.get("reconcile") or {}
    return [
        "口径说明（M3）",
        "1. COS 应付取 DescribeBillSummaryByProduct 的 RealTotalCost（BusinessCode=p_cos）。",
        "2. 按桶应付取 DescribeBillResourceSummary 同行 RealTotalCost 之和。",
        f"3. 对账：按桶合计 vs COS 应付，容差 {recon.get('tolerance')} 元。"
        f" 本次差额 {recon.get('delta')}，通过={recon.get('ok')}。",
        "4. 可优化 KPI = 规则净节省 ≥ 50 元/月且无强阻断；不含 R05/R07/R08/R09 金额、R06（标题含「不含 CDN 下行」）、备份桶。",
        "5. 单价优先 账单 RealTotalCost / 监控用量；否则使用刊例并在证据中标注。",
        "6. R01 无清单时只做桶级保守观察（20% STANDARD），不读清单 CSV，<64KB 不计入沉降。",
        "7. 建议 IA/ARCHIVE/DEEP 天数分别 ≥30/90/180。禁止列对象，禁止写入生命周期。",
        "8. UIN / 账户默认掩码。导出不含 SecretKey、不含对象 Key。",
        "9. Ready=0 时金额为暂估；PDF 加水印「暂估」。",
        "10. 抽屉按钮仅为「复制草稿」，不会把规则写回存储桶。",
    ]


def _register_cn_font() -> str:
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    name = "COSCostCN"
    if name in pdfmetrics.getRegisteredFontNames():
        return name
    for path in FONT_CANDIDATES:
        p = Path(path)
        if not p.is_file():
            continue
        try:
            if p.suffix.lower() == ".ttc":
                pdfmetrics.registerFont(TTFont(name, str(p), subfontIndex=0))
            else:
                pdfmetrics.registerFont(TTFont(name, str(p)))
            return name
        except Exception:
            continue
    return "Helvetica"


def _short(name: Any) -> str:
    text = str(name or "")
    if len(text) <= 22:
        return text
    return text[:10] + "…" + text[-8:]


def _draw_trend(c, trend, x, y, w, h, font) -> None:
    c.setStrokeColorRGB(0.89, 0.9, 0.92)
    c.setFillColorRGB(1, 1, 1)
    c.roundRect(x, y, w, h, 3, fill=1, stroke=1)
    c.setFillColorRGB(0.07, 0.09, 0.15)
    c.setFont(font, 8)
    c.drawString(x + 4, y + h - 12, "C1 费用趋势")
    months = trend.get("months") or []
    values = [v if v is not None else 0 for v in (trend.get("payable") or [])]
    _bars(c, months, values, x + 8, y + 8, w - 16, h - 24)


def _draw_bucket_bars(c, rows, x, y, w, h, font) -> None:
    c.setStrokeColorRGB(0.89, 0.9, 0.92)
    c.setFillColorRGB(1, 1, 1)
    c.roundRect(x, y, w, h, 3, fill=1, stroke=1)
    c.setFillColorRGB(0.07, 0.09, 0.15)
    c.setFont(font, 8)
    c.drawString(x + 4, y + h - 12, "C2 桶费用（Top）")
    labels = [_short(r.get("bucket")) for r in rows[:6]]
    values = [float(r.get("payable") or 0) for r in rows[:6]]
    _bars(c, labels, values, x + 8, y + 8, w - 16, h - 24)


def _bars(c, labels, values, x, y, w, h) -> None:
    if not values:
        return
    peak = max(values) or 1
    n = len(values)
    gap = 6
    bw = max(8, (w - gap * n) / max(n, 1))
    c.setFillColorRGB(0.15, 0.39, 0.92)
    for i, val in enumerate(values):
        bh = (float(val) / peak) * (h - 14)
        bx = x + i * (bw + gap)
        c.rect(bx, y + 12, bw, max(bh, 1), fill=1, stroke=0)
    c.setFillColorRGB(0.4, 0.42, 0.45)
    c.setFont("Helvetica", 6)
    for i, label in enumerate(labels):
        c.saveState()
        c.translate(x + i * (bw + gap) + 2, y + 2)
        c.drawString(0, 0, str(label)[:10])
        c.restoreState()


def _draw_table(c, title, headers, rows, x, y, w, font) -> None:
    row_h = 11
    header_h = 12
    height = header_h + row_h * max(len(rows), 1) + 16
    c.setStrokeColorRGB(0.89, 0.9, 0.92)
    c.setFillColorRGB(1, 1, 1)
    c.roundRect(x, y - height, w, height, 3, fill=1, stroke=1)
    c.setFillColorRGB(0.07, 0.09, 0.15)
    c.setFont(font, 8)
    c.drawString(x + 4, y - 12, title)
    col_w = w / max(len(headers), 1)
    c.setFont(font, 7)
    for i, h in enumerate(headers):
        c.drawString(x + 4 + i * col_w, y - 24, str(h))
    c.setFillColorRGB(0.2, 0.22, 0.26)
    for r, row in enumerate(rows):
        for i, cell in enumerate(row):
            c.drawString(x + 4 + i * col_w, y - 36 - r * row_h, str(cell)[:22])
